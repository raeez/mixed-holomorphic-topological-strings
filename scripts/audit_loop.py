#!/Users/raeez/mixed-holomorphic-topological-strings/scripts/.venv-audit/bin/python
"""Audit loop for main23_platonic_upgrade_audit.xlsx.

Reads the page-grounded upgrade audit (1,718 entries) into a canonical
JSON form, tracks per-entry resolution state, and orders work along the
monotone spine T_finite < T_coord < ... < T_compact under the
no-downgrade rule from CLAUDE.md/AGENTS.md and the audit's Executive_Spine.

Subcommands:
  import   refresh _audit/audit.json from the XLSX, seed _audit/state.json
  list     filter entries by category/severity/theorem-level/status/page
  show     print one entry's full record (audit + state)
  mark     record status / verification / commit / note for one entry
  next     propose the next K entries to address, monotone-spine ordered
  progress total + per-level + per-severity status counts
  monotone print the monotone-spine table + check status monotonicity
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import sys
from pathlib import Path

ROOT = Path("/Users/raeez/mixed-holomorphic-topological-strings")
AUDIT_DIR = ROOT / "_audit"
AUDIT_JSON = AUDIT_DIR / "audit.json"
STATE_JSON = AUDIT_DIR / "state.json"

# Default XLSX source (override on `import`).
DEFAULT_XLSX = Path("/Users/raeez/Downloads/main23_platonic_upgrade_audit.xlsx")

# Monotone-spine rank.  T_spine (cross-level structural) ranks first as
# foundational infrastructure; the canonical spine T_finite -> T_compact
# follows; T_exposition ranks last.  T_adm and T_quant/radial slot at the
# admissible-topology and quantization rungs respectively.
LEVEL_RANK = {
    "T_spine":           0,
    "T_finite":          1,
    "T_coord":           2,
    "T_P0 / T_coord":    2,
    "T_P0":              3,
    "T_adm":             4,
    "T_kernel":          5,
    "T_quant/radial":    6,
    "T_bar/cobar":       7,
    "T_Hoch":            8,
    "T_Hoch / T_trace":  8,
    "T_Capelli":         9,
    "T_QME":            10,
    "T_KN/global":      11,
    "T_compact":        12,
    "T_exposition":     99,
}

SEVERITY_RANK = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

VALID_STATUSES = (
    "actionable",
    "in_progress",
    "resolved",
    "verified",
    "deferred",
    "blocked",
)


def _utcnow() -> str:
    return _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# import: XLSX -> audit.json, seed state.json
# ---------------------------------------------------------------------------

def cmd_import(args: argparse.Namespace) -> int:
    import openpyxl  # local import; only needed for import path

    xlsx = Path(args.xlsx).expanduser().resolve()
    if not xlsx.exists():
        print(f"error: {xlsx} not found", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, object]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[sheet_name] = []
            continue
        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        records: list[dict[str, object]] = []
        for raw in rows[1:]:
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
                continue
            rec = {header[i]: raw[i] for i in range(len(header))}
            records.append(rec)
        sheets[sheet_name] = records

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": str(xlsx),
        "imported_at": _utcnow(),
        "sheets": sheets,
    }
    AUDIT_JSON.write_text(json.dumps(payload, indent=2, default=str))
    print(f"wrote {AUDIT_JSON}")
    for s, rows in sheets.items():
        print(f"  {s}: {len(rows)} rows")

    # Seed state.json only if absent; never overwrite resolution progress.
    if STATE_JSON.exists() and not args.reset_state:
        print(f"keeping existing {STATE_JSON} (pass --reset-state to wipe)")
    else:
        entries = sheets.get("Audit_1000_plus", [])
        state = {}
        for rec in entries:
            entry_id = rec.get("ID")
            if entry_id is None:
                continue
            key = str(entry_id)
            raw_status = (rec.get("Status") or "").strip().lower()
            status = raw_status if raw_status in VALID_STATUSES else "actionable"
            state[key] = {
                "status": status,
                "updated_at": _utcnow(),
                "commit": None,
                "verification": None,
                "notes": None,
            }
        STATE_JSON.write_text(json.dumps(state, indent=2))
        print(f"seeded {STATE_JSON} with {len(state)} entries")
    return 0


# ---------------------------------------------------------------------------
# Loading helpers
# ---------------------------------------------------------------------------

def _load_audit() -> dict:
    if not AUDIT_JSON.exists():
        print(f"error: {AUDIT_JSON} not found; run `audit_loop.py import` first",
              file=sys.stderr)
        sys.exit(2)
    return json.loads(AUDIT_JSON.read_text())


def _load_state() -> dict:
    if not STATE_JSON.exists():
        return {}
    return json.loads(STATE_JSON.read_text())


def _save_state(state: dict) -> None:
    STATE_JSON.write_text(json.dumps(state, indent=2))


def _entries() -> list[dict[str, object]]:
    return _load_audit()["sheets"]["Audit_1000_plus"]


def _entry_priority(entry: dict[str, object], status: str) -> tuple:
    level = str(entry.get("Theorem level") or "")
    sev = str(entry.get("Severity") or "")
    page = entry.get("Page")
    try:
        page_n = int(page)
    except (TypeError, ValueError):
        page_n = 10**6
    # Lower tuple sorts first.  Resolved/verified rank after open work.
    status_rank = 0 if status in ("actionable", "in_progress", "blocked") else 1
    return (
        status_rank,
        LEVEL_RANK.get(level, 50),
        SEVERITY_RANK.get(sev, 9),
        page_n,
        int(entry.get("ID") or 0),
    )


# ---------------------------------------------------------------------------
# list
# ---------------------------------------------------------------------------

def _filter_entries(entries, state, args) -> list[dict[str, object]]:
    out = []
    for e in entries:
        if args.category and args.category.lower() not in str(e.get("Category", "")).lower():
            continue
        if args.severity and args.severity.lower() != str(e.get("Severity", "")).lower():
            continue
        if args.theorem_level and args.theorem_level.lower() not in str(e.get("Theorem level", "")).lower():
            continue
        if args.finding_type and args.finding_type.lower() not in str(e.get("Finding type", "")).lower():
            continue
        if args.page is not None and int(e.get("Page") or -1) != args.page:
            continue
        if args.status:
            cur = state.get(str(e.get("ID")), {}).get("status", "actionable")
            if cur != args.status:
                continue
        out.append(e)
    return out


def cmd_list(args: argparse.Namespace) -> int:
    entries = _entries()
    state = _load_state()
    filtered = _filter_entries(entries, state, args)
    filtered.sort(key=lambda e: _entry_priority(e, state.get(str(e.get("ID")), {}).get("status", "actionable")))
    if not filtered:
        print("(no entries match)")
        return 0
    limit = args.limit if args.limit > 0 else len(filtered)
    print(f"matched {len(filtered)} entries; showing {min(limit, len(filtered))}")
    print(f"{'ID':>5}  {'p':>3}  {'level':<18}  {'sev':<8}  {'type':<18}  {'status':<11}  category :: issue")
    print("-" * 140)
    for e in filtered[:limit]:
        eid = str(e.get("ID"))
        st = state.get(eid, {}).get("status", "actionable")
        issue = (str(e.get("Issue / technical minutia / improvement") or "")).strip().replace("\n", " ")
        print(
            f"{eid:>5}  {str(e.get('Page') or ''):>3}  "
            f"{(str(e.get('Theorem level') or ''))[:18]:<18}  "
            f"{(str(e.get('Severity') or ''))[:8]:<8}  "
            f"{(str(e.get('Finding type') or ''))[:18]:<18}  "
            f"{st:<11}  "
            f"{(str(e.get('Category') or ''))[:32]} :: {issue[:60]}"
        )
    return 0


# ---------------------------------------------------------------------------
# show
# ---------------------------------------------------------------------------

def cmd_show(args: argparse.Namespace) -> int:
    target = str(args.id)
    entries = _entries()
    state = _load_state()
    for e in entries:
        if str(e.get("ID")) == target:
            print(f"=== Entry {target} ===")
            for k, v in e.items():
                if v is None or v == "":
                    continue
                if isinstance(v, str) and len(v) > 200:
                    print(f"{k}:")
                    print(f"  {v}")
                else:
                    print(f"{k}: {v}")
            st = state.get(target, {})
            print("\n--- resolution state ---")
            for k, v in st.items():
                print(f"{k}: {v}")
            return 0
    print(f"error: entry {target} not found", file=sys.stderr)
    return 1


# ---------------------------------------------------------------------------
# mark
# ---------------------------------------------------------------------------

def cmd_mark(args: argparse.Namespace) -> int:
    target = str(args.id)
    entries = _entries()
    found = any(str(e.get("ID")) == target for e in entries)
    if not found:
        print(f"error: entry {target} not found in audit.json", file=sys.stderr)
        return 1
    state = _load_state()
    rec = state.get(target, {
        "status": "actionable",
        "updated_at": _utcnow(),
        "commit": None,
        "verification": None,
        "notes": None,
    })
    if args.status:
        if args.status not in VALID_STATUSES:
            print(f"error: status must be one of {VALID_STATUSES}", file=sys.stderr)
            return 2
        rec["status"] = args.status
    if args.commit is not None:
        rec["commit"] = args.commit
    if args.verification is not None:
        rec["verification"] = args.verification
    if args.note is not None:
        rec["notes"] = args.note
    rec["updated_at"] = _utcnow()
    state[target] = rec
    _save_state(state)
    print(f"updated entry {target}: {rec}")
    return 0


# ---------------------------------------------------------------------------
# next
# ---------------------------------------------------------------------------

def cmd_next(args: argparse.Namespace) -> int:
    entries = _entries()
    state = _load_state()
    open_entries = []
    for e in entries:
        eid = str(e.get("ID"))
        st = state.get(eid, {}).get("status", "actionable")
        if st in ("actionable", "in_progress", "blocked"):
            open_entries.append((e, st))
    open_entries.sort(key=lambda pair: _entry_priority(pair[0], pair[1]))
    if not open_entries:
        print("(no open entries — audit fully resolved)")
        return 0
    limit = args.limit if args.limit > 0 else 5
    print(f"next {min(limit, len(open_entries))} of {len(open_entries)} open entries (monotone-spine ordered):\n")
    for e, st in open_entries[:limit]:
        eid = str(e.get("ID"))
        print(f"ID {eid}  page {e.get('Page')}  {e.get('Theorem level')}  [{e.get('Severity')}]  {e.get('Finding type')}  status={st}")
        print(f"  category: {e.get('Category')}")
        print(f"  issue: {e.get('Issue / technical minutia / improvement')}")
        print(f"  upgrade: {e.get('Strongest upgrade construction')}")
        print(f"  probe: {e.get('Verification / consistency probe')}")
        print()
    return 0


# ---------------------------------------------------------------------------
# bulk-mark
# ---------------------------------------------------------------------------

def cmd_bulk_mark(args: argparse.Namespace) -> int:
    entries = _entries()
    state = _load_state()

    class _F:
        pass
    flt = _F()
    flt.category = args.category
    flt.severity = args.severity
    flt.theorem_level = args.theorem_level
    flt.finding_type = args.finding_type
    flt.status = args.from_status
    flt.page = None
    matched = _filter_entries(entries, state, flt)
    if not matched:
        print("(no entries match the filter)")
        return 0
    now = _utcnow()
    touched = 0
    for e in matched:
        key = str(e.get("ID"))
        rec = state.get(key, {
            "status": "actionable",
            "updated_at": now,
            "commit": None,
            "verification": None,
            "notes": None,
        })
        rec["status"] = args.status
        if args.verification is not None:
            rec["verification"] = args.verification
        if args.note is not None:
            rec["notes"] = args.note
        if args.commit is not None:
            rec["commit"] = args.commit
        rec["updated_at"] = now
        state[key] = rec
        touched += 1
    if args.dry_run:
        print(f"DRY RUN: would touch {touched} entries -> status={args.status}")
        return 0
    _save_state(state)
    print(f"bulk-marked {touched} entries -> status={args.status}")
    if args.verification:
        print(f"  verification: {args.verification}")
    if args.note:
        print(f"  notes: {args.note}")
    return 0


# ---------------------------------------------------------------------------
# progress
# ---------------------------------------------------------------------------

def cmd_progress(args: argparse.Namespace) -> int:
    entries = _entries()
    state = _load_state()
    by_status: dict[str, int] = {}
    by_level_status: dict[tuple, int] = {}
    by_severity_status: dict[tuple, int] = {}
    for e in entries:
        eid = str(e.get("ID"))
        st = state.get(eid, {}).get("status", "actionable")
        by_status[st] = by_status.get(st, 0) + 1
        lvl = str(e.get("Theorem level") or "")
        sev = str(e.get("Severity") or "")
        by_level_status[(lvl, st)] = by_level_status.get((lvl, st), 0) + 1
        by_severity_status[(sev, st)] = by_severity_status.get((sev, st), 0) + 1
    total = len(entries)
    print(f"total entries: {total}\n")
    print("by status:")
    for st in VALID_STATUSES:
        n = by_status.get(st, 0)
        if n:
            print(f"  {st:<12}  {n:5d}  ({100*n/total:5.1f}%)")
    print("\nby theorem level x status:")
    levels = sorted({lvl for (lvl, _) in by_level_status}, key=lambda l: LEVEL_RANK.get(l, 50))
    for lvl in levels:
        line = [f"  {lvl:<20}"]
        for st in VALID_STATUSES:
            n = by_level_status.get((lvl, st), 0)
            line.append(f"{st[:4]}={n:4d}")
        print(" ".join(line))
    print("\nby severity x status:")
    for sev in ("Critical", "High", "Medium", "Low"):
        line = [f"  {sev:<10}"]
        for st in VALID_STATUSES:
            n = by_severity_status.get((sev, st), 0)
            line.append(f"{st[:4]}={n:4d}")
        print(" ".join(line))
    return 0


# ---------------------------------------------------------------------------
# monotone
# ---------------------------------------------------------------------------

def cmd_monotone(args: argparse.Namespace) -> int:
    audit = _load_audit()
    spine = audit["sheets"].get("Monotone_Spine", [])
    print("Monotone spine (XLSX):\n")
    for row in spine:
        lvl = row.get("Level") or ""
        strong = row.get("Strong object") or ""
        shadow = row.get("Forgetful shadow") or ""
        witness = row.get("Strictness witness") or ""
        print(f"  {lvl:<15}  {strong[:54]}  shadow={shadow[:36]}  witness={witness[:36]}")
    print()
    # Sanity check: every level that appears in Audit_1000_plus has a known rank.
    entries = audit["sheets"].get("Audit_1000_plus", [])
    seen = sorted({str(e.get("Theorem level") or "") for e in entries})
    print("Theorem levels seen in audit (with rank):")
    for lvl in seen:
        rank = LEVEL_RANK.get(lvl)
        flag = "" if rank is not None else "  <-- UNRANKED"
        print(f"  {lvl:<24}  rank={rank if rank is not None else '?':<5}{flag}")
    return 0


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(prog="audit_loop", description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd", required=True)

    pi = sub.add_parser("import", help="refresh audit.json from XLSX, seed state.json")
    pi.add_argument("--xlsx", default=str(DEFAULT_XLSX),
                    help=f"XLSX path (default: {DEFAULT_XLSX})")
    pi.add_argument("--reset-state", action="store_true",
                    help="overwrite existing state.json (destroys resolution progress)")

    pl = sub.add_parser("list", help="list entries with optional filters")
    for arg in ("category", "severity", "theorem-level", "finding-type", "status"):
        pl.add_argument(f"--{arg}", default=None)
    pl.add_argument("--page", type=int, default=None)
    pl.add_argument("--limit", type=int, default=50)

    ps = sub.add_parser("show", help="show one entry's full record")
    ps.add_argument("id")

    pm = sub.add_parser("mark", help="record state for one entry")
    pm.add_argument("id")
    pm.add_argument("--status", choices=VALID_STATUSES, default=None)
    pm.add_argument("--commit", default=None, help="git commit hash that addressed this entry")
    pm.add_argument("--verification", default=None, help="verification probe outcome")
    pm.add_argument("--note", default=None, help="free-text note")

    pn = sub.add_parser("next", help="propose next K entries to address")
    pn.add_argument("--limit", type=int, default=5)

    pb = sub.add_parser("bulk-mark",
                        help="apply uniform state to many entries by filter")
    for arg in ("category", "severity", "theorem-level", "finding-type"):
        pb.add_argument(f"--{arg}", default=None)
    pb.add_argument("--from-status", default=None,
                    help="only touch entries whose current status equals this")
    pb.add_argument("--status", choices=VALID_STATUSES, required=True)
    pb.add_argument("--verification", default=None)
    pb.add_argument("--note", default=None)
    pb.add_argument("--commit", default=None)
    pb.add_argument("--dry-run", action="store_true")

    sub.add_parser("progress", help="summary of resolution progress")
    sub.add_parser("monotone", help="print the monotone-spine table")

    args = p.parse_args(argv)
    return {
        "import":    cmd_import,
        "list":      cmd_list,
        "show":      cmd_show,
        "mark":      cmd_mark,
        "next":      cmd_next,
        "bulk-mark": cmd_bulk_mark,
        "progress":  cmd_progress,
        "monotone":  cmd_monotone,
    }[args.cmd](args)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
